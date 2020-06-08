
###############################################################################################
###############################################################################################

# Name:             Random_Fallow_Field_Selection.py
# Author:           Kelly Meehan, USBR
# Created:          20191104
# Updated:          20200608 
# Version:          Created using Python 3.6.8 

# Requires:         Pyhton 3.x

# Notes:            This script is intended to be used as a stand-alone script.

# Description:      This script allows a user to randomly select fields participating in a fallowing program for on the ground visual inspection.
                  
#----------------------------------------------------------------------------------------------

# This script will:

# 0. Set-up
# 1. Read in csv and clean data
# 2. Create new data frame from slice of master data frame 
# 3. Collapse rows by field_id (i.e. remove duplicates) but combine acreage
# 4. Randomly sort rows
# 5. Loop to rotate through unique sections and selecting fields belonging to unchosen ones so long as acreage doesn't exceed target
# 6. Create a subset of original data frame with only those rows corresponding to selected fields
# 7. Write data frame to original Excel file 

###############################################################################################
###############################################################################################

# 0. Set-up

# 0.0 Install necessary packages
import os, pandas

#----------------------------------------------------------------------------------------------

# 1. Read in csv and clean data

# Create pandas data frame from spreadsheet, excluding first 6 rows
input_spreadsheet = r'C:\Temp\Random_Field_Selection.xlxs'
df_pvid = pandas.read_excel(input_spreadsheet, sheet_name = 'USBR INSP FALLOW 12-2019 ', skiprows = 5, converters = {'Unnamed: 2':str, ' # South':str, '# East': str, 'Unnamed: 10':int})

# Rename columns
df_pvid.columns = ['field_id', 'farm_id', 'section', 'township', 'range', 'acct_id', 'parcel_id', 'qualified_acres', 'parcel_id', 'canal_gate', 'fallowed_acreage', 'special_location', 'zip_code', 'blank', 'date_fallowed', 'duration_fallowed']

# Remove null rows and erroneous sum row
df_pvid = df_pvid[pandas.notnull(df_pvid['farm_id'])]

# Calculate total acreage
total_acreage = df_pvid['fallowed_acreage'].sum()

# Calculate target acreage at 5% of total acreage
target_acreage = total_acreage * 0.05

# Run multiple iterations of random selection of fields until a selection meeting the desired acreage and number of fields as produced

number_fields_selected = 0
acreage_selected = 0.0

#while number_fields_selected == 0 or number_fields_selected > 30 or number_fields_selected < 25 or acreage_selected < target_acreage:
while number_fields_selected == 0 or number_fields_selected > 30 or number_fields_selected < 25:

    # 2. Create new data frame from slice of master data frame 
    df_random = df_pvid[['field_id', 'section', 'fallowed_acreage']]
    
    # # 3. If you wish to collapse rows by field_id (i.e. remove duplicates) but combine acreage, uncomment this section
    # df_random = df_random.groupby(['field_id'], as_index = False).agg({'section': 'first', 'fallowed_acreage': 'sum'})
    
    # 4. Randomly sort rows
    df_random = df_random.sample(frac = 1)
    
    # 5. Loop to rotate through unique sections and selecting fields belonging to unchosen ones so long as acreage doesn't exceed target
    
    # Generate list of unique sections (unique zip codes in region; i.e. every township/range combo has one unique section)
    unique_sections = df_random['section'].unique()
    
    # Set counter for selected acreage
    acreage_selected = 0.0
    
    # Create empty lists for field_id and section so as to never select duplicates
    fields_selected = []
    sections_selected = []
    
    # Iterate over sections and then over rows of data frame, selecting rows so long as neither the field_id nor section of the field being iterated through has been chosen yet 
    for s in unique_sections:
        for i, row in df_random.iterrows():
            if acreage_selected < target_acreage:
                if df_random.at[i, 'section'] == s and df_random.at[i, 'field_id'] not in fields_selected and df_random.at[i, 'section'] not in sections_selected and df_random.at[i, 'fallowed_acreage'] >= 5:
                    acreage_selected += df_random.at[i, 'fallowed_acreage']
                    fields_selected.append(df_random.at[i, 'field_id'])
                    sections_selected.append(df_random.at[i, 'section'])
                
    print(acreage_selected)
    print(fields_selected)
    print(sections_selected)
    number_fields_selected = len(fields_selected)
    print(number_fields_selected)

# 6. Create a subset of original data frame with only those rows corresponding to selected fields
df_selected = df_pvid[df_pvid['field_id'].isin(fields_selected)]
df_selected

# 8. Reset index of data frame to field_id
df_selected.set_index('field_id', inplace = True)
df_selected

# 7. Write data frame to original Excel file 

# Import function, load_workbook, from openpyxl which allows you to add a sheet to an already existing Excel file
from openpyxl import load_workbook 

# Check to see if Excel file already exists; if so, add just a sheet; otherwise, create the Excel file
if os.path.isfile(input_spreadsheet):
    book = load_workbook(input_spreadsheet)
    writer = pandas.ExcelWriter(input_spreadsheet, engine ='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    df_selected.to_excel(writer, 'Selected_Fields')
    writer.save()

else:
    df_selected.to_excel(excel_writer = r'C:\Temp\Random_Field_Selection.xlxs', sheet_name = 'Selected_Fields')

